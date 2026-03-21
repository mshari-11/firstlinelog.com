
import json, os, io, re, logging
from datetime import datetime, timezone
from decimal import Decimal, ROUND_HALF_UP
import boto3
from botocore.exceptions import ClientError
import openpyxl
from openpyxl.styles import Alignment

logger = logging.getLogger()
logger.setLevel(logging.INFO)

REGION           = os.environ.get("AWS_DEFAULT_REGION", "me-south-1")
TEMPLATES_BUCKET = os.environ["FINANCE_TEMPLATES_BUCKET"]
EXPORTS_BUCKET   = os.environ["FINANCE_EXPORTS_BUCKET"]
TEMPLATE_KEY     = "stc_template.xlsx"
PRESIGNED_EXPIRY = 3600
s3 = boto3.client("s3", region_name=REGION)

def normalize_phone(raw):
    digits = re.sub(r'\D', '', str(raw))
    if len(digits) == 9 and digits.startswith('5'):
        return '966' + digits
    elif len(digits) == 10 and digits.startswith('05'):
        return '966' + digits[1:]
    elif len(digits) == 12 and digits.startswith('966'):
        return digits
    raise ValueError(f"Invalid STC phone: {raw!r} -> digits={digits!r}")

def to_decimal(v):
    return Decimal(str(v)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

def load_template():
    resp = s3.get_object(Bucket=TEMPLATES_BUCKET, Key=TEMPLATE_KEY)
    return openpyxl.load_workbook(io.BytesIO(resp["Body"].read()), data_only=False)

def find_data_start(ws):
    for r in range(1, 10):
        v = ws.cell(row=r, column=1).value
        if v is None or str(v).strip() == "":
            return r
    return 2

def write_payouts(ws, payouts, start_row):
    total, written = Decimal("0.00"), 0
    for i, p in enumerate(payouts):
        row = start_row + i
        ref   = str(p.get("reference","")).strip()
        phone = str(p.get("phone","")).strip()
        amt   = p.get("amount")
        if not ref:     raise ValueError(f"Row {row}: reference required")
        if not phone:   raise ValueError(f"Row {row}: phone required")
        if amt is None: raise ValueError(f"Row {row}: amount required")
        norm_phone = normalize_phone(phone)
        amount = to_decimal(amt)
        if amount <= 0: raise ValueError(f"Row {row}: amount must be > 0")
        ca = ws.cell(row=row, column=1, value=ref)
        ca.number_format = "@"; ca.alignment = Alignment(horizontal="left")
        cb = ws.cell(row=row, column=2, value=norm_phone)
        cb.number_format = "@"; cb.alignment = Alignment(horizontal="left")
        cc = ws.cell(row=row, column=3, value=float(amount))
        cc.number_format = "#,##0.00"; cc.alignment = Alignment(horizontal="right")
        total += amount; written += 1
    return written, total

def upload_wb(wb, s3_key):
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    s3.put_object(Bucket=EXPORTS_BUCKET, Key=s3_key, Body=buf.getvalue(),
        ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ServerSideEncryption="AES256")

def presign(s3_key):
    fname = s3_key.split("/")[-1]
    return s3.generate_presigned_url("get_object",
        Params={"Bucket": EXPORTS_BUCKET, "Key": s3_key,
                "ResponseContentDisposition": f'attachment; filename="{fname}"'},
        ExpiresIn=PRESIGNED_EXPIRY)

def _error(code, msg):
    return {"statusCode": code,
            "headers": {"Content-Type": "application/json"},
            "body": json.dumps({"status": "error", "message": msg}, ensure_ascii=False)}

def lambda_handler(event, context):
    body = event
    if isinstance(event.get("body"), str):
        body = json.loads(event["body"])
    elif isinstance(event.get("body"), dict):
        body = event["body"]
    try:
        run_id  = str(body.get("run_id", "UNKNOWN")).strip()
        period  = str(body.get("period", "")).strip()
        payouts = body.get("payouts", [])
        if not payouts: return _error(400, "No payouts provided")
        if len(payouts) > 5000: return _error(400, f"Too many rows: {len(payouts)}")
        ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        filename = f"STC_{run_id}_{ts}.xlsx"
        pfolder  = period or datetime.now(timezone.utc).strftime("%Y-%m")
        s3_key   = f"exports/{pfolder}/{filename}"
        logger.info(f"Generating {filename} | rows={len(payouts)} | run={run_id}")
        wb = load_template(); ws = wb.active
        start = find_data_start(ws)
        row_count, total = write_payouts(ws, payouts, start)
        upload_wb(wb, s3_key)
        url = presign(s3_key)
        result = {"status": "ok", "filename": filename, "s3_key": s3_key,
                  "presigned_url": url, "row_count": row_count,
                  "total_amount": float(total), "period": pfolder,
                  "run_id": run_id,
                  "generated_at": datetime.now(timezone.utc).isoformat()}
        logger.info(f"SUCCESS: {row_count} rows | total={total} SAR")
        return {"statusCode": 200,
                "headers": {"Content-Type": "application/json",
                            "Access-Control-Allow-Origin": "https://staff.fll.sa"},
                "body": json.dumps(result, ensure_ascii=False)}
    except ValueError as e:
        return _error(422, str(e))
    except ClientError as e:
        logger.error(f"AWS error: {e}")
        return _error(500, f"AWS: {e.response['Error']['Code']}")
    except Exception as e:
        logger.exception("Unexpected error")
        return _error(500, str(e))
